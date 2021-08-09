import time
import urllib.request

# image_url = 'https://bit.ly/2XuVzB4' #the image on the web
# save_name = './img/my_image.jpg' #local name to be saved
# urllib.request.urlretrieve(image_url, save_name)


# import module
import openpyxl
import urllib.request, urllib.error
import io
from urllib.request import urlopen, Request
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

# load excel with its path
# wrkbk = openpyxl.load_workbook("full-list.json.xlsx") // BEFORE!!!
wrkbk = openpyxl.load_workbook("smart-braslet.xlsx")

sh = wrkbk.active

# iterate through excel and display data
for i in range(1, sh.max_row + 1):
    time.sleep(0.5)
    print("\n")
    # print("Row ", i, " data :")

    # only first mcolumn
    for j in range(4, 5):
        # column 1 (cell_obj.value = value of cell)
        cell_obj = sh.cell(row=i, column=j)
        # Name of image
        save_name = './img/wathc__' + str(i) + '.jpg'
        print(save_name)
        # Save image
        # urllib.request.urlretrieve(cell_obj.value, save_name)

        # Try for errors
        url = cell_obj.value
        try:
            conn = urllib.request.urlopen(url)
        except urllib.error.HTTPError as e:
            # Save just text file if there is error
            f = open('./img/wathc__' + str(i) + 'remove.txt', 'w+')
            f.write('Just for test')
            f.close()
            print('HTTPError: {}'.format(e.code))
        except urllib.error.URLError as e:
            # Save just text file if there is error
            f = open('./img/wathc__' + str(i) + 'remove.txt', 'w+')
            f.write('Just for test')
            f.close()
            print('URLError: {}'.format(e.reason))
        else:
            # Save image
            urllib.request.urlretrieve(cell_obj.value, save_name)