# import urllib.request

# image_url = 'https://bit.ly/2XuVzB4' #the image on the web
# save_name = './img/my_image.jpg' #local name to be saved
# urllib.request.urlretrieve(image_url, save_name)




# import module
import openpyxl
import urllib.request, urllib.error
import io

  
# load excel with its path
# wrkbk = openpyxl.load_workbook("ForParcing.xlsx") // BEFORE!!!
wrkbk = openpyxl.load_workbook("ForParcing.xlsx")
  
sh = wrkbk.active
  
# iterate through excel and display data
for i in range(1, sh.max_row+1):
    print("\n")
    # print("Row ", i, " data :")
      
    # only first mcolumn
    for j in range(5, 6):
        # column 1 (cell_obj.value = value of cell)
        cell_obj = sh.cell(row=i, column=j)
        # Name of image
        save_name = './img/product__' + str(i) + '.jpg'
        print(save_name)
        # Save image
        # urllib.request.urlretrieve(cell_obj.value, save_name)   

        # Try for errors
        url = cell_obj.value
        try:
            conn = urllib.request.urlopen(url)
        except urllib.error.HTTPError as e:
            # Save just text file if there is error
            f= open('./img/product__' + str(i) + 'remove.txt','w+')
            f.write('Just for test')
            f.close() 
            print('HTTPError: {}'.format(e.code))
        except urllib.error.URLError as e:
            # Save just text file if there is error
            f= open('./img/product__' + str(i) + 'remove.txt','w+')
            f.write('Just for test')
            f.close() 
            print('URLError: {}'.format(e.reason))
        else:
            # Save image
            urllib.request.urlretrieve(cell_obj.value, save_name)   

        

